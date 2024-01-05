from openpyxl import Workbook

# Create a new instance of a workbook
# There will be one empty worksheet in it (default)
wb = Workbook()

# Get the reference of the worksheet at index 0
ws1 = wb.worksheets[0]
# Set a new title
ws1.title = "First Sheet"
# Assign a value to cell A1
ws1["A1"].value = "Test"

# Create an additional work sheet with a title of "Second Sheet"
# The new worksheet will be added to the end by default
# create_sheet(title, position)
wb.create_sheet("Second Sheet")

# Get the reference of the worksheet by its title
ws2 = wb["Second Sheet"]

# Assign a value to cell B3 (in worksheet "Second Sheet")
ws2.cell(row=3, column=2, value="New Value")  # ws2["B3"].value = "New Value"
# Assign an Excel formula to cell C5 (in worksheet "Second Sheet")
ws2["C5"].value = "=SUM(3, 5)"

# Save the workbook
wb.save("wb1.xlsx")
