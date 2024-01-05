import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter import messagebox as msg
from openpyxl import load_workbook
from openpyxl import Workbook


def add_new(event):
    new_entry = [txt_fname.get(), txt_lname.get(), txt_grade.get()]
    tv.insert(parent="", index="end", values=new_entry)
    tv.yview_moveto(1.0)
    txt_fname.delete(first=0, last="end")
    txt_lname.delete(first=0, last="end")
    txt_grade.delete(first=0, last="end")
    txt_fname.focus_set()


def import_data():
    file_filter = (("Excel files", "*.xlsx"),
                   ("All files", "*.*"))
    filename = fd.askopenfile(title="Choose file", filetypes=file_filter)

    if filename is not None:
        wb = load_workbook(filename.name)
        ws1 = wb.active
        data = []
        skip_first = True
        for row in ws1.iter_rows():
            if skip_first:
                skip_first = False
                continue
            for cell in row:
                data.append(cell.value)
            tv.insert(parent="", index="end", values=data)
            data.clear()


def export_data():
    if len(tv.get_children()) == 0:
        msg.showinfo(title="Export Data", message="No data available to export.")
        return

    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="First Name")
    ws.cell(row=1, column=2, value="Last Name")
    ws.cell(row=1, column=3, value="Grade")

    row_count = 2
    for x in tv.get_children():
        # print(tv.item(x)["values"])
        ws.cell(row=row_count, column=1, value=tv.item(x)["values"][0])
        ws.cell(row=row_count, column=2, value=tv.item(x)["values"][1])
        ws.cell(row=row_count, column=3, value=tv.item(x)["values"][2])
        row_count += 1

    ws.cell(row=row_count, column=3, value=f"=AVERAGE(C2:C{(row_count-1)})")

    wb.save("exported.xlsx")
    msg.showinfo(title="Export Data", message="The data has been exported and saved as 'exported.xlsx'.")


win = tk.Tk()
win.geometry("420x200+700+200")
win.title("SEN4017 - Week 13")
win.resizable(False, False)

top_container = tk.Frame(win)
bottom_container = tk.Frame(win)
top_container.pack(fill="both", expand=True)
bottom_container.pack(fill="both", expand=True)

tv = ttk.Treeview(top_container, show="headings", height=5)
tv["columns"] = ("fname", "lname", "grade")
tv.pack(fill="both", expand=True, padx=(10, 25), pady=15)

tv.heading("fname", text="First Name", anchor="center")
tv.heading("lname", text="Last Name", anchor="center")
tv.heading("grade", text="Grade", anchor="center")

tv.column("fname", anchor="center", width=150)
tv.column("lname", anchor="center", width=150)
tv.column("grade", anchor="center", width=70)

tv_scroll = ttk.Scrollbar(top_container, orient="vertical", command=tv.yview)
tv.configure(yscrollcommand=tv_scroll.set)
tv_scroll.place(relx=1, rely=0, relheight=1, anchor="ne")

txt_fname = ttk.Entry(bottom_container, width=24)
txt_lname = ttk.Entry(bottom_container, width=24)
txt_grade = ttk.Entry(bottom_container, width=10)

txt_fname.grid(row=0, column=0, padx=10)
txt_lname.grid(row=0, column=1)
txt_grade.grid(row=0, column=2, padx=10)

txt_grade.bind("<Return>", add_new)

menubar = tk.Menu(win)
win.configure(menu=menubar)

file_menu = tk.Menu(menubar, tearoff=False)
file_menu.add_command(label="Import Data...", command=import_data)
file_menu.add_command(label="Export Data", command=export_data)
menubar.add_cascade(label="File", menu=file_menu)

win.mainloop()
